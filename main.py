import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import openpyxl


# Apertura del libro de excel + seleccion de hoja principal
db = openpyxl.load_workbook('dbrh.xlsx')
hoja = db['Candidatos']        

#####################################
#           FUNCIONES               #
#####################################


def guardarNuevoCandidato():     
    # Igualacion de variables a los entry del formulario                                               
    nombre = caja_nombre.get()
    apellido = caja_apellido.get()
    edad = caja_edad.get()
    email = caja_mail.get()
    grupo = lista_grupo.get()
    conocimientos = []

    # Asignacion de fila y ID a utilizar teniendo en cuenta el último dígito usado (almacenado en el .txt) 
    fileFila = open("ultimaFila.txt", "r")                                             
    ultimaFila = fileFila.readline().strip()
    ultimaFilaInt = int(ultimaFila)                                                 
    fila = ultimaFilaInt+1 
    fileId = open("ultimoID.txt", "r")
    id = fileId.readline().strip()
    idInt = int(id)                                                             
    
    # Escritura de los datos del entry en el excel
    hoja.cell(row=fila, column=1, value=idInt) # Almaceno ID
    hoja.cell(row=fila, column=2, value=nombre)
    hoja.cell(row=fila, column=3, value=apellido)
    hoja.cell(row=fila, column=4, value=edad)
    hoja.cell(row=fila, column=5, value=email)
    hoja.cell(row=fila, column=6, value=grupo)
    
    # Sobreescribir el archivo txt con la ultima fila utilizada
    fileFila = open('ultimaFila.txt', 'w')                                              
    fileFila.write(str(fila))                                                                                                
    fileFila.close()

    # Sobreescribir el archivo txt con el ultimo ID utilizado
    fileId = open('ultimoID.txt', 'w')
    nuevoId = idInt+1                                              
    fileId.write(str(nuevoId))                                                                                                
    fileId.close()

    # Guardado del excel
    db.save('dbrh.xlsx')

    # Reset de los entry + Msg de confirmación
    caja_nombre.delete(0,"end")
    caja_apellido.delete(0,"end")
    caja_edad.delete(0,"end")
    caja_mail.delete(0,"end")
    messagebox.showinfo('Confirmacion', 'El candidato ha sido agregado correctamente.')






###################################
#           VENTANA               #
###################################

ventana = tk.Tk()                                                               
ventana.config(width=400, height=500)                                           
ventana.title('Administrador de RRHH')



#####################################
#    VENTANA AGREGAR CANDIDATOS     #
#####################################

label_add = tk.Label(text='Agregar candidatos:')
label_add.place(x=20, y=20, width=120, height=30)

# Nombre
label_nombre_candidato = tk.Label(text='Nombre:')
label_nombre_candidato.place(x=15, y=55, width=70, height=30)
caja_nombre = tk.Entry()
caja_nombre.place(x=90, y=55, width=80, height=25)

# Apellido
label_apellido_candidato = tk.Label(text='Apellido:')
label_apellido_candidato.place(x=15, y=90, width=70, height=30)
caja_apellido = tk.Entry()
caja_apellido.place(x=90, y=90, width=80, height=25)

# Edad
label_edad_candidato = tk.Label(text='Edad:')
label_edad_candidato.place(x=15, y=125, width=70, height=30)
caja_edad = tk.Entry()
caja_edad.place(x=90, y=125, width=80, height=25)

# Mail
label_mail_candidato = tk.Label(text='Mail:')
label_mail_candidato.place(x=15, y=160, width=70, height=30)
caja_mail = tk.Entry()
caja_mail.place(x=90, y=160, width=80, height=25)

# Lista desplegable 'Funcion'
label_lista_grupo = tk.Label(text='Función:')
label_lista_grupo.place(x=15, y=195, width=70, height=30)
lista_grupo = ttk.Combobox(state='readonly', 
                        values=['Front End Developer', 'Back End Developer', 'QA Tester', 'Full Stack', 'Diseñador gráfico', 'Otro'] )
lista_grupo.place(x=90, y=200, width=125, height=20)

opcionSeleccionada = lista_grupo.get()

# Boton de confirmacion
boton_confirmarCandidato = tk.Button(text='Aceptar', command=guardarNuevoCandidato)
boton_confirmarCandidato.place(x=50, y=260, width=100, height=30)


ventana.mainloop()                                                              # Muestro la ventana en loop para que se actualice constantemente.
